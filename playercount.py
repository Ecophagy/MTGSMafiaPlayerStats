#
# Reads the files in Input/, strips player number indicators, 
# extracts the first word of the player name and writes them into an excel document
#

import re
import openpyxl
import os.path


def import_raw_player_data():

    wb = openpyxl.load_workbook('Analysis.xlsx')
    ws = wb["rawdata"] # Get sheet

    current_col = 1

    for year in range(2010, 2021):
        # Write column Header
        _ = ws.cell(column=current_col, row=1, value=year)
        for quarter in range(1,5):
            # Write column header
            _ = ws.cell(column=current_col, row=2, value="Q"+str(quarter))

            # Ingest raw input
            file = os.path.join("Input", "MafiaPlayerList" + str(year) + "Q" + str(quarter) + ".txt")
            # Check if file exists (for years where not all data is available)
            if os.path.isfile(file):
                with open(file, 'r') as f:

                    # Remove unused lines
                    lines = (line.rstrip("\r\n''") for line in f)

                    full_list = list()

                    # Remove lines which are blank
                    for ln in lines:
                        if ln == "":
                            pass
                        else:
                            full_list.append(ln)

                    name_list = list()

                    # Remove formatting!
                    for line in full_list:
                        # Remove numbers and punctuation before player TODO: Remove "The"?
                        line = re.sub('^[0-9]+[\.\)]*', '', line)
                        # Remove commas, semicolons, colons, asterisks,tags,dashes,brackets,full stops
                        line = re.sub('[,;:\*\(\).]', '', line)
                        # Remove any lingering bbcode tags
                        line = re.sub('\[.*\]', '', line)
                        # set to all lowercase
                        line = line.lower()
                        # Remove "replacement" indicators
                        line = re.sub('^(replacing|r\.*|re|rep|rr|replaced)', '', line)
                        # Take only the first word
                        name_list.append(line.split()[0])

                    name_seen = dict()

                    # Remove duplicates and count them
                    # TODO: Use the counts for something
                    for name in name_list:
                        if name not in name_seen:
                            name_seen[name] = 1
                        else:
                            name_seen[name] = name_seen[name] + 1

                    # Write list out to excel
                    current_row = 3  # Start below headers
                    for name in name_seen:
                        _ = ws.cell(column=current_col, row = current_row, value = name)
                        current_row += 1

                    # Go to next column
                    current_col += 1

                    wb.save(filename="Analysis.xlsx")


if __name__ == "__main__":
    import_raw_player_data()