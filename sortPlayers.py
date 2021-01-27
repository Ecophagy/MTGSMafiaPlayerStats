#
# Reads player information out of an excel document, 
# calculates the sizes of the various populations, and then writes these into the excel document.
#

import openpyxl


def sort_players():
    wb = openpyxl.load_workbook('Analysis.xlsx')
    ws = wb["rawdata"]  # Get Sheet

    players = []

    for i in range(1, 45):  # Number of columns in rawdata sheet
        templist=[]
        current_row = 3  # Skip headers
        while True:
            #  Go down the column until we hit an empty cell (None)
            val = ws.cell(column=i, row=current_row).value
            if val is not None:
                templist.append(val)
                current_row += 1
            else:
                players.append(templist)
                break

    new_players = []
    active_players = []
    returning_players = []
    all_players = []

    for i in range(0, len(players)):
        new_player_count = 0
        active_player_count = 0
        returning_player_count = 0
        for j in range(0, len(players[i])):
            # new player check
            if players[i][j] not in all_players:
                new_player_count += 1
                all_players.append(players[i][j])
            else:
                # check for returning players (last game was > 6 months ago)
                if i > 2 and players[i][j] not in players[i-1] and players[i][j] not in players[i-2]:
                    returning_player_count += 1
                else:
                    active_player_count += 1

        new_players.append(new_player_count)
        active_players.append(active_player_count)
        returning_players.append(returning_player_count )

    ws = wb["playergroups"]
    # Set headers
    _=ws.cell(column=1, row=2, value="New")
    _=ws.cell(column=1, row=3, value="Active")
    _=ws.cell(column=1, row=4, value="Returning")
    current_column = 2
    for year in range(2010, 2021):
        for quarter in range(1,5):
            _ = ws.cell(column=current_column, row=1, value=str(year) + "Q" + str(quarter))
            current_column += 1

    # populate!
    current_column = 2
    for number in new_players:
        _ = ws.cell(column=current_column, row=2, value=number)
        current_column += 1

    current_column = 2
    for number in active_players:
        _ = ws.cell(column=current_column, row=3, value=number)
        current_column += 1

    current_column = 2
    for number in returning_players:
        _ = ws.cell(column=current_column, row=4, value=number)
        current_column += 1

    wb.save(filename="Analysis.xlsx")


if __name__ == "__main__":
    sort_players()